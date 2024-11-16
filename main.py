import os
import re
from datetime import datetime
from telethon.sync import TelegramClient
from telethon.tl.functions.channels import GetParticipantsRequest
from telethon.tl.types import ChannelParticipantsSearch
from telethon.errors import FloodWaitError
import openpyxl
import time
from config import API_ID, API_HASH, PHONE
import asyncio

class TeleParser:
    def __init__(self):
        self.base_dir = "parsed_channels"
        self._ensure_base_dir()
        self.filter_chars = (
            list('абвгдеёжзийклмнопрстуфхцчшщъыьэюя') +
            list('abcdefghijklmnopqrstuvwxyz') +
            list('0123456789')
        )
    
    def _ensure_base_dir(self):
        if not os.path.exists(self.base_dir):
            os.makedirs(self.base_dir)
    
    def _sanitize_filename(self, filename):
        sanitized = re.sub(r'[<>:"/\\|?*]', '', filename)
        sanitized = sanitized.replace(' ', '_')
        return sanitized
    
    def _create_channel_dir(self, channel_name):
        safe_channel_name = self._sanitize_filename(channel_name)
        channel_dir = os.path.join(self.base_dir, safe_channel_name)
        if not os.path.exists(channel_dir):
            os.makedirs(channel_dir)
        return channel_dir
    
    async def get_channel_members(self, channel_link):
        async with TelegramClient('session_name', API_ID, API_HASH) as client:
            await client.start(phone=PHONE)
            
            try:
                channel = await client.get_entity(channel_link)
                channel_name = channel.title
                all_participants = []
                seen_ids = set()  # Множество для хранения ID пользователей
                
                print(f"\nПарсинг канала: {channel_name}")
                print("Сбор участников...")
                
                total_count = (await client.get_participants(channel, limit=0)).total
                print(f"Всего участников в канале: {total_count}")
                
                for char in self.filter_chars:
                    offset = 0
                    limit = 200
                    
                    while True:
                        try:
                            participants = await client(GetParticipantsRequest(
                                channel,
                                ChannelParticipantsSearch(char),
                                offset,
                                limit,
                                hash=0
                            ))
                            
                            if not participants.users:
                                break
                            
                            # Добавляем только уникальных пользователей по ID
                            new_count = 0
                            for user in participants.users:
                                if user.id not in seen_ids:
                                    seen_ids.add(user.id)
                                    all_participants.append(user)
                                    new_count += 1
                            
                            if new_count > 0:
                                print(f"Собрано участников: {len(all_participants)} (поиск по символу '{char}')")
                            
                            if len(participants.users) < limit:
                                break
                            
                            offset += limit
                            await asyncio.sleep(2)
                            
                        except FloodWaitError as e:
                            print(f"Достигнут лимит запросов. Ожидание {e.seconds} секунд...")
                            await asyncio.sleep(e.seconds)
                            continue
                        except Exception as e:
                            print(f"Ошибка при поиске по символу '{char}': {str(e)}")
                            break
                
                print(f"\nИтого собрано уникальных участников: {len(all_participants)}")
                return channel_name, all_participants
                
            except Exception as e:
                print(f"Ошибка: {str(e)}")
                return None, []

    def save_to_excel(self, channel_name, participants):
        channel_dir = self._create_channel_dir(channel_name)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"members_{timestamp}.xlsx"
        filepath = os.path.join(channel_dir, filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Участники канала"
        
        headers = [
            "Username", "Имя", "Фамилия", 
            "ID пользователя", "Премиум статус",
            "Био", "Онлайн статус"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for idx, user in enumerate(participants, 2):
            ws.cell(row=idx, column=1, value=user.username or "—")
            ws.cell(row=idx, column=2, value=user.first_name or "—")
            ws.cell(row=idx, column=3, value=user.last_name or "—")
            ws.cell(row=idx, column=4, value=str(user.id))
            ws.cell(row=idx, column=5, value="Да" if user.premium else "Нет")
            ws.cell(row=idx, column=7, value="Недавно онлайн" if user.status else "—")
        
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        print(f"\nДанные сохранены в файл: {filepath}")
        return filepath

def main():
    print("╔════════════════════════════════════╗")
    print("║    TeleParser | @marcelosacred     ║")
    print("║    Парсер участников Telegram      ║")
    print("╚════════════════════════════════════╝\n")
    
    parser = TeleParser()
    
    while True:
        channel_link = input("\nВведите ссылку на канал (или 'q' для выхода): ")
        
        if channel_link.lower() == 'q':
            break
            
        channel_link = channel_link.replace("https://t.me/", "")
        channel_link = channel_link.replace("@", "")
        
        channel_name, participants = asyncio.get_event_loop().run_until_complete(
            parser.get_channel_members(channel_link)
        )
        
        if channel_name and participants:
            print(f"\nНайдено участников: {len(participants)}")
            parser.save_to_excel(channel_name, participants)
        else:
            print("Участники не найдены или произошла ошибка")
        
        print("\n" + "="*50)

if __name__ == "__main__":
    main()